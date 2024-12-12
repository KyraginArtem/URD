# client/views/template_constructor_window.py
# ========== Импорты ==========
import json
import os
from email.policy import strict

from PySide6 import QtGui
from PySide6.QtGui import QIcon
from PySide6.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QComboBox, QPushButton, QTableWidget,
    QHBoxLayout, QTableWidgetItem, QMessageBox, QFontComboBox, QColorDialog, QLineEdit, QDateEdit
)
from PySide6.QtCore import Qt, Signal, QFile, QSize, QDate
from openpyxl.descriptors import DateTime

from client.services.table_cell_parser import TableCellParser
from client.services.template_table_service import TemplateTableService
from client.views.window_creating_new_template import WindowCreatingNewTemplate

# ========== Класс TemplateConstructorWindow ==========
class TemplateConstructorWindow(QWidget):
    # -------- Сигналы --------
    create_report = Signal()
    create_requested = Signal(str)  # Открытие окна создания нового шаблона
    template_selected = Signal(str)  # Выбор шаблона
    template_save_requested = Signal()  # Сохранение шаблона
    delete_template_signal = Signal(str)  # Удаление шаблона
    template_update = Signal(str, int, int, str)  # Обновление шаблона

    # -------- Инициализация --------
    def __init__(self, admin_name, template_names, controller):
        super().__init__()
        self.decimal_display = None
        self.setWindowTitle("Редактор шаблонов")
        self.resize(1000, 700)

        # Параметры
        self.admin_name = admin_name
        self.template_names = template_names
        self.controller = controller

        # Переменные интерфейса
        self.start_date = None
        self.end_date = None
        self.template_combo = None
        self.template_name = ""
        self.settings_window = None
        self.template_name_label = None
        self.font_combo_box = None
        self.background_color = "#FFFFFF"

        # Инициализация интерфейса
        self.init_ui()
        self.load_styles()

    # -------- Основной метод создания интерфейса --------
    def init_ui(self):
        """
        Формирует весь интерфейс: панель, меню, таблицу.
        """
        # Основной лейаут
        main_layout = QVBoxLayout(self)

        # Создание и добавление частей интерфейса
        main_layout.addLayout(self.create_top_panel())  # Верхняя панель
        lower_menu = self.create_side_menus()  # нижнее меню меню
        main_layout.addLayout(lower_menu)
        self.create_table(main_layout)  # Таблица

        # Установка лейаута
        self.setLayout(main_layout)

    # -------- Методы создания секций интерфейса --------
    def create_top_panel(self):
        """
        Создает верхнюю панель с выпадающим меню, кнопками, метками.
        """
        layout = QHBoxLayout()

        # Выпадающее меню для выбора шаблона
        self.template_combo = QComboBox()
        self.template_combo.addItems(self.template_names)
        self.template_combo.currentTextChanged.connect(self.name_on_template_selected)
        self.template_combo.currentTextChanged.connect(
            lambda: self.template_selected.emit(self.template_combo.currentText())
        )
        layout.addWidget(self.template_combo)

        # Кнопка для создания нового шаблона
        create_button = QPushButton("Создать шаблон")
        create_button.clicked.connect(lambda: self.create_requested.emit(self.template_name))
        layout.addWidget(create_button)

        # Метка имени шаблона
        self.template_name_label = QLabel("Шаблон: Не выбран")
        layout.addWidget(self.template_name_label)

        # Поля для выбора даты
        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)  # Показываем календарь
        self.start_date.setDisplayFormat("dd.MM.yyyy")  # Формат отображения даты
        self.start_date.setDate(QDate(2024, 1, 1))
        #self.start_date.setDate(QDate.currentDate().addDays(-7))  # Устанавливаем начальную дату по умолчанию
        layout.addWidget(QLabel("С даты:"))
        layout.addWidget(self.start_date)

        self.end_date = QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDisplayFormat("dd.MM.yyyy")
        # self.end_date.setDate(QDate.currentDate())  # Устанавливаем текущую дату по умолчанию
        self.end_date.setDate(QDate(2024, 1, 10))
        layout.addWidget(QLabel("По дату:"))
        layout.addWidget(self.end_date)

        # Кнопка для формирования отчета
        generate_report_button = QPushButton("Сформировать отчет")
        generate_report_button.clicked.connect(self.create_report)  # Подключаем обработчик кнопки
        layout.addWidget(generate_report_button)

        # Метка имени администратора
        admin_label = QLabel(self.admin_name)
        admin_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        layout.addWidget(admin_label)

        return layout

    def create_side_menus(self):
        current_dir = os.path.dirname(os.path.abspath(__file__))

        # Нижнее меню
        lower_menu = QHBoxLayout()

        # Кнопка "Объединить"
        merge_button = QPushButton()
        merge_button.setIcon(QIcon(os.path.join(current_dir, "..", "icons", "merge_cells.png")))
        merge_button.setIconSize(QSize(24, 24))
        merge_button.setFixedSize(24, 24)
        merge_button.clicked.connect(self.handle_merge_cells)
        lower_menu.addWidget(merge_button)

        #Выделенный текст
        bold_button = self.create_icon_button("icon_bold_text.png",
                                              lambda: TemplateTableService.toggle_bold(self.table))
        lower_menu.addWidget(bold_button)

        #наклон текста
        italic_button = self.create_icon_button("icon_italic_text.png",
                                              lambda: TemplateTableService.toggle_italic(self.table))
        lower_menu.addWidget(italic_button)

        #подчеркивание
        underline_button = self.create_icon_button("icon_underline_text.png",
                                                lambda: TemplateTableService.toggle_underline(self.table))
        lower_menu.addWidget(underline_button)

        #цвет текста
        text_color = self.create_icon_button("icon_color_text.png", lambda: self.change_text_color())
        lower_menu.addWidget(text_color)

        #цвет ячейки
        cell_color = self.create_icon_button("icon_color_cell.png", lambda:  self.change_cell_color())
        lower_menu.addWidget(cell_color)

        #цвет фона
        template_background_color = self.create_icon_button("icon_background.png",
                                                            lambda: self.change_template_background_color())
        lower_menu.addWidget(template_background_color)

        # Кнопка для увеличения количества знаков после запятой
        increase_decimal = self.create_icon_button(
            "icon_increase_decimal.png",
            lambda: self.handle_decimal_change(increase=True)
        )
        lower_menu.addWidget(increase_decimal)

        # Кнопка для уменьшения количества знаков после запятой
        decrease_decimal = self.create_icon_button(
            "icon_decrease_decimal.png",
            lambda: self.handle_decimal_change(increase=False)
        )
        lower_menu.addWidget(decrease_decimal)

        # Поле для отображения текущего количества знаков после запятой
        self.decimal_display = QLineEdit()
        self.decimal_display.setFixedSize(38, 24)
        self.decimal_display.setReadOnly(True)  # Делаем поле некликабельным
        self.decimal_display.setAlignment(Qt.AlignCenter)
        lower_menu.addWidget(self.decimal_display)

        #Сдвиг
        #----------------------------------
        # В лево
        shift_left_button = self.create_icon_button("free-icon-arrow-left.png",
                                                    lambda: self.shift_selection("left"))
        lower_menu.addWidget(shift_left_button)

        # В право
        shift_right_button = self.create_icon_button("free-icon-arrow-right.png",
                                                     lambda: self.shift_selection("right"))
        lower_menu.addWidget(shift_right_button)

        # В верх
        shift_up_button = self.create_icon_button("free-icon-arrow-up.png",
                                                  lambda: self.shift_selection("up"))
        lower_menu.addWidget(shift_up_button)

        # В низ
        shift_down_button = self.create_icon_button("free-icon-arrow-down.png",
                                                    lambda: self.shift_selection("down"))
        lower_menu.addWidget(shift_down_button)
        #----------------------------------

        #Выбор шрифта
        self.font_combo_box = QFontComboBox()
        self.font_combo_box.setFixedWidth(150)
        self.font_combo_box.currentFontChanged.connect(self.change_font)
        lower_menu.addWidget(self.font_combo_box)

        # Добавляем выпадающий список для выбора размера шрифта
        font_size_combo_box = QComboBox()
        font_size_combo_box.setFixedSize(40, 24)  # Устанавливаем размер
        font_size_combo_box.addItems([str(size) for size in range(8, 73, 2)])  # Размеры шрифтов от 8 до 72 с шагом 2
        font_size_combo_box.currentTextChanged.connect(
            lambda size: TemplateTableService.change_font_size(self.table, int(size)))
        lower_menu.addWidget(font_size_combo_box)

        # Поле для ввода количества строк
        row_input = QLineEdit()
        row_input.setFixedWidth(30)  # Устанавливаем ширину поля
        row_input.setValidator(QtGui.QIntValidator(1, 99))  # Ограничиваем ввод только числами от 1 до 1000
        lower_menu.addWidget(row_input)

        # Поле для ввода количества столбцов
        column_input = QLineEdit()
        column_input.setFixedWidth(30)
        column_input.setValidator(QtGui.QIntValidator(1, 99))
        lower_menu.addWidget(column_input)

        #Кнопка применить изменения размера
        resize_button = self.create_icon_button("icon_check.png",
                                            lambda: TemplateTableService.resize_table(self.table,
                                            int(row_input.text()), int(column_input.text()), self.background_color))
        lower_menu.addWidget(resize_button)

        download_file = self.create_icon_button("icon_download.png",
                                            lambda: TemplateTableService().export_to_excel(self.table))
        lower_menu.addWidget(download_file)

        save_button = self.create_icon_button("icon_save.png", self.template_save_requested.emit)
        lower_menu.addWidget(save_button)

        delete_button = self.create_icon_button("icon_delete.png", self.delete_template)
        lower_menu.addWidget(delete_button)
        lower_menu.setSpacing(10)

        return lower_menu

    def create_table(self, layout):
        """
        Создает таблицу и добавляет её в лейаут.
        """
        self.table = QTableWidget(0, 0)
        self.table.setHorizontalHeaderLabels([chr(65 + i) for i in range(4)])
        self.table.setVerticalHeaderLabels([str(i + 1) for i in range(4)])
        # self.table.itemChanged.connect(lambda item: self.handle_cell_change(item.row(), item.column()))
        layout.addWidget(self.table)

        #self.table.itemSelectionChanged.connect(self.update_decimal_display)

    def create_icon_button(self, icon_name, callback):
        """
        Создает кнопку с иконкой.
        """
        button = QPushButton()
        button.setIcon(QIcon(os.path.join(os.path.dirname(__file__), "..", "icons", icon_name)))
        button.setIconSize(QSize(24, 24))
        button.setFixedSize(24, 24)
        button.clicked.connect(callback)
        return button

    # -------- Загрузка стилей --------
    def load_styles(self):
        """
        Загружает стили из файла .qss.
        """
        current_dir = os.path.dirname(os.path.abspath(__file__))
        style_file_path = os.path.join(current_dir, "..", "styles", "styles.qss")
        file = QFile(style_file_path)
        if file.open(QFile.ReadOnly | QFile.Text):
            stylesheet = file.readAll().data().decode()
            self.setStyleSheet(stylesheet)
            print("Стили успешно загружены")
        else:
            print(f"Ошибка: не удалось открыть файл стилей по пути: {style_file_path}")

    def load_styles(self):
        """Загружает стили из файла .qss"""
        current_dir = os.path.dirname(os.path.abspath(__file__))
        style_file_path = os.path.join(current_dir, "..", "styles", "styles.qss")
        file = QFile(style_file_path)
        if file.open(QFile.ReadOnly | QFile.Text):
            stylesheet = file.readAll().data().decode()
            self.setStyleSheet(stylesheet)
            print("Стили успешно загружены")
        else:
            print(f"Ошибка: не удалось открыть файл стилей по пути: {style_file_path}")

    def change_font(self, font):
        # Получаем список выделенных ячеек
        selected_ranges = self.table.selectedRanges()

        for selected_range in selected_ranges:
            for row in range(selected_range.topRow(), selected_range.bottomRow() + 1):
                for col in range(selected_range.leftColumn(), selected_range.rightColumn() + 1):
                    item = self.table.item(row, col)
                    if item is None:
                        item = QTableWidgetItem()
                        self.table.setItem(row, col, item)

                    # Применяем новый шрифт к ячейке
                    item.setFont(font)

                    # Сохраняем данные о шрифте в Qt.UserRole
                    cell_data_str = item.data(Qt.UserRole)
                    if cell_data_str:
                        cell_data = json.loads(cell_data_str)
                    else:
                        cell_data = {}

                    # Обновляем данные о шрифте
                    cell_data['font'] = font.toString()
                    item.setData(Qt.UserRole, json.dumps(cell_data))

                    # Обновляем отображение
                    item.setTextAlignment(Qt.AlignCenter)

    # Выбор имени в комбо-боксе
    def name_on_template_selected(self, template_name):
        self.template_name = template_name
        self.template_name_label.setText(f"Шаблон: {template_name}")

    #Перед отрисовкой нового шаблона необходимо обновить структуру шаблона
    def update_table_structure(self, row_count, col_count, column_labels):
        # Сброс объединений ячеек перед загрузкой нового шаблона
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                self.table.setSpan(row, col, 1, 1)  # Сбрасываем объединение для каждой ячейки

        # Очищаем содержимое таблицы от данных
        self.table.clearContents()
        self.table.setRowCount(row_count)
        self.table.setColumnCount(col_count)
        # Обращается к сервису для генерации цифровых и буквенных имен ячеек
        self.table.setHorizontalHeaderLabels(column_labels)
        self.table.setVerticalHeaderLabels([str(i + 1) for i in range(row_count)])

    def update_template_name(self, template_name):
        # Обновляем имя шаблона в метке
        self.template_name = template_name
        self.template_name_label.setText(f"Шаблон: {template_name}")

    def add_template_name_to_combo(self, template_name):
    # Проверяем, что имя шаблона не пустое и не дублируется
        if template_name and template_name not in [self.template_combo.itemText(i) for i in range(self.template_combo.count())]:
            print(f"Добавляем шаблон '{template_name}' в комбо бокс.")
            self.template_combo.addItem(template_name)
        else:
            print(f"Шаблон '{template_name}' уже существует в комбо боксе или имя пустое.")

    def remove_template_name_from_combo(self, template_name):
        index = self.template_combo.findText(template_name)
        if index != -1:
            self.template_combo.removeItem(index)

    #Метод закомментирован для оптимизации
    def set_current_template_in_combo(self, template_name):
        # Устанавливаем текущий элемент в комбо-боксе на только что добавленный шаблон
        index = self.template_combo.findText(template_name)
        if index != -1:
            self.template_combo.setCurrentIndex(index)

    def delete_template(self):
        if not self.template_name:
            QMessageBox.warning(self, "Ошибка удаления", "Пожалуйста, выберите шаблон для удаления.")
            return
        confirm = QMessageBox.question(
            self, "Подтверждение удаления", f"Вы уверены, что хотите удалить шаблон '{self.template_name}'?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if confirm == QMessageBox.Yes:
            self.delete_template_signal.emit(self.template_name)

    def handle_merge_cells(self):
        selected_ranges = self.table.selectedRanges()
        if len(selected_ranges) == 1:
            merge_range = selected_ranges[0]
            top_row = merge_range.topRow()
            bottom_row = merge_range.bottomRow()
            left_col = merge_range.leftColumn()
            right_col = merge_range.rightColumn()

            # Проверяем, является ли верхняя левая ячейка объединенной
            if TemplateTableService.is_merged(self.table, top_row, left_col):
                # Если ячейки объединены, отменяем объединение
                self.controller.handle_unmerge_cells_request(top_row, bottom_row, left_col, right_col)
            else:
                # Если ячейки не объединены, выполняем объединение
                self.controller.handle_merge_cells_request(top_row, bottom_row, left_col, right_col)

    def change_text_color(self):
        """Открывает палитру для выбора цвета текста и применяет его к выделенным ячейкам."""
        color = QColorDialog.getColor()
        if color.isValid():
            TemplateTableService.apply_text_color(self.table, color.name())

    def change_cell_color(self):
        """Открывает палитру для выбора цвета фона ячейки и применяет его к выделенным ячейкам."""
        color = QColorDialog.getColor()
        if color.isValid():
            TemplateTableService.apply_cell_background_color(self.table, color.name())

    def change_template_background_color(self):
        """Открывает палитру для изменения общего цвета фона шаблона."""
        color = QColorDialog.getColor()
        if color.isValid():
            self.background_color = color.name()
            self.table.setStyleSheet(f"background-color: {self.background_color};")


    def update_background_color(self, color):
        """Обновляет цвет фона таблицы."""
        self.background_color = color  # Сохраняем цвет в переменную
        self.table.setStyleSheet(f"background-color: {color};")

    def handle_decimal_change(self, increase):
        """
        Изменяет количество знаков после запятой и обновляет отображение.
        """
        TemplateTableService.change_decimal_places(self.table, increase)
        self.update_decimal_display()

    def update_decimal_display(self):
        """Обновляет отображение текущего количества знаков после запятой."""
        selected_ranges = self.table.selectedRanges()
        if not selected_ranges:
            self.decimal_display.setText("")
            return

        # Берем первую выделенную ячейку
        selected_range = selected_ranges[0]
        row, col = selected_range.topRow(), selected_range.leftColumn()
        item = self.table.item(row, col)
        if item is None:
            self.decimal_display.setText("0")
            return

        # Получаем текущий формат из конфигурации
        cell_data_str = item.data(Qt.UserRole)
        if cell_data_str:
            cell_config = json.loads(cell_data_str)
            current_format = cell_config.get("format", 0)
        else:
            current_format = 0

        self.decimal_display.setText(str(current_format))

    def shift_selection(self, direction):
        """Сдвигает выделенные ячейки в указанном направлении."""
        selected_ranges = self.table.selectedRanges()
        if not selected_ranges:
            return

        for selected_range in selected_ranges:
            top_row, bottom_row = selected_range.topRow(), selected_range.bottomRow()
            left_col, right_col = selected_range.leftColumn(), selected_range.rightColumn()

            # Сохраняем исходные данные и конфигурации ячеек
            original_data = []
            for row in range(top_row, bottom_row + 1):
                row_data = []
                for col in range(left_col, right_col + 1):
                    item = self.table.item(row, col)
                    row_data.append({
                        "value": item.text() if item else "",
                        "config": json.loads(item.data(Qt.UserRole)) if item and item.data(Qt.UserRole) else {}
                    })
                original_data.append(row_data)

            # Переносим данные из соседней области в область выделения
            if direction == "left":
                for row in range(top_row, bottom_row + 1):
                    for col in range(left_col, right_col + 1):
                        source_col = col - 1
                        if source_col >= 0:
                            self.copy_cell_data(row, col, row, source_col)
                        else:
                            self.clear_cell(row, col)
            elif direction == "right":
                for row in range(top_row, bottom_row + 1):
                    for col in range(right_col, left_col - 1, -1):
                        source_col = col + 1
                        if source_col < self.table.columnCount():
                            self.copy_cell_data(row, col, row, source_col)
                        else:
                            self.clear_cell(row, col)
            elif direction == "up":
                for col in range(left_col, right_col + 1):
                    for row in range(top_row, bottom_row + 1):
                        source_row = row - 1
                        if source_row >= 0:
                            self.copy_cell_data(row, col, source_row, col)
                        else:
                            self.clear_cell(row, col)
            elif direction == "down":
                for col in range(left_col, right_col + 1):
                    for row in range(bottom_row, top_row - 1, -1):
                        source_row = row + 1
                        if source_row < self.table.rowCount():
                            self.copy_cell_data(row, col, source_row, col)
                        else:
                            self.clear_cell(row, col)

            # Переносим исходные данные в новое место и восстанавливаем конфигурации
            for r, row_data in enumerate(original_data):
                for c, cell_data in enumerate(row_data):
                    target_row = top_row + r
                    target_col = left_col + c
                    if direction == "left":
                        target_col -= 1
                    elif direction == "right":
                        target_col += 1
                    elif direction == "up":
                        target_row -= 1
                    elif direction == "down":
                        target_row += 1

                    if 0 <= target_row < self.table.rowCount() and 0 <= target_col < self.table.columnCount():
                        # Создаём ячейку и задаём значение
                        item = QTableWidgetItem(cell_data["value"])
                        self.table.setItem(target_row, target_col, item)

                        # Применяем конфигурации
                        if cell_data["config"]:
                            TemplateTableService.apply_cell_configuration(
                                item, self.table, target_row, target_col, cell_data["config"]
                            )
                            # Сохраняем конфигурацию в Qt.UserRole
                            item.setData(Qt.UserRole, json.dumps(cell_data["config"]))

                        # Применяем центровку текста
                        item.setTextAlignment(Qt.AlignCenter)

    def copy_cell_data(self, target_row, target_col, source_row, source_col):
        """Копирует все данные и конфигурации из одной ячейки в другую."""
        source_item = self.table.item(source_row, source_col)
        if source_item:
            # Создаём копию текста и данных ячейки
            target_item = QTableWidgetItem(source_item.text())
            target_item.setData(Qt.UserRole, source_item.data(Qt.UserRole))

            # Копируем выравнивание текста
            target_item.setTextAlignment(source_item.textAlignment())

            # Копируем фон ячейки
            target_item.setBackground(source_item.background())

            # Копируем шрифт
            target_item.setFont(source_item.font())

            # Копируем цвет текста
            target_item.setForeground(source_item.foreground())
        else:
            # Если ячейка пуста, создаём новую
            target_item = QTableWidgetItem("")

        # Устанавливаем скопированную ячейку на целевую позицию
        self.table.setItem(target_row, target_col, target_item)

    def clear_cell(self, row, col):
        """Очищает содержимое ячейки."""
        self.table.setItem(row, col, QTableWidgetItem(""))


