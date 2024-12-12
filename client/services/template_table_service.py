# client/services/template_table_service.py
import json

from PySide6 import QtGui
from PySide6.QtGui import Qt, QColor
from PySide6.QtWidgets import QTableWidgetItem, QColorDialog, QFileDialog, QMessageBox
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


# from client.services.abstract_table_service import AbstractTableService

class TemplateTableService:

    @staticmethod
    def merge_cells(table, top_row, bottom_row, left_col, right_col, main_value):
        merge_range = f"{TemplateTableService.generate_cell_name(top_row, left_col)}:" \
                      f"{TemplateTableService.generate_cell_name(bottom_row, right_col)}"
        table.setSpan(top_row, left_col, (bottom_row - top_row + 1), (right_col - left_col + 1))
        cell_data = {
            "merger": merge_range
        }
        item = QTableWidgetItem()
        TemplateTableService.set_cell_data(item, main_value, cell_data)  # Используем функцию для установки данных
        table.setItem(top_row, left_col, item)

        # Очищаем остальные ячейки в диапазоне
        for row in range(top_row, bottom_row + 1):
            for col in range(left_col, right_col + 1):
                if not (row == top_row and col == left_col):
                    table.setItem(row, col, None)  # Очистить ячейку, так как она теперь часть объединения

    @staticmethod
    def column_label_to_index(label):
        """Преобразует буквенный заголовок столбца (например, 'A', 'B', 'AA') в индекс."""
        label = label.upper()
        index = 0
        for i, char in enumerate(reversed(label)):
            index += (ord(char) - ord('A') + 1) * (26 ** i)
        return index - 1

    #Генерация цифровых и буквенных имен ячеек в зависимости от кол-ва
    @staticmethod
    def generate_cell_name(row, col):
        labels = []
        num_columns = col + 1
        while num_columns > 0:
            num_columns -= 1
            labels.append(chr(65 + (num_columns % 26)))
            num_columns //= 26

        row_label = str(row + 1)
        return f"{''.join(reversed(labels))}{row_label}"

    @staticmethod
    def generate_col_name(col):
        labels = []
        num_columns = col + 1
        while num_columns > 0:
            num_columns -= 1
            labels.append(chr(65 + (num_columns % 26)))
            num_columns //= 26
        return f"{''.join(reversed(labels))}"

    @staticmethod
    def collect_table_data(table_widget):
        """Собирает данные из таблицы и возвращает их в формате JSON."""
        rows = table_widget.rowCount()
        cols = table_widget.columnCount()
        cell_data = []

        for row in range(rows):
            for col in range(cols):
                item = table_widget.item(row, col)
                cell_value = ""

                # Генерируем имя ячейки (например, A1, B2)
                cell_name = TemplateTableService.generate_cell_name(row, col)
                cell_config = {
                    "background_color": None,
                    "height": None,
                    "width": None,
                    "text_color": None,
                    "font": None,
                    "format": None,
                    "text_tilt": None,
                    "underline": None,
                    "text_size": None,
                    "cell_name": cell_name,
                    "merger": None,
                    "bold": None
                }

                if item:
                    # Получаем текст из ячейки
                    cell_value = item.text()

                    # Получаем данные из Qt.UserRole, если они существуют
                    cell_data_str = item.data(Qt.UserRole)
                    if cell_data_str:
                        try:
                            # Пытаемся разобрать строку данных как JSON
                            cell_data_json = json.loads(cell_data_str)
                            #Записываем ширину и высоту
                            cell_data_json['height'] =  table_widget.rowHeight(row)
                            cell_data_json['width'] = table_widget.columnWidth(col)
                            # Обновляем конфигурацию из данных в Qt.UserRole
                            for key in cell_config.keys():
                                if key in cell_data_json:
                                    cell_config[key] = cell_data_json[key]
                        except json.JSONDecodeError:
                            print(f"Ошибка парсинга JSON в ячейке ({row}, {col}): {cell_data_str}")

                # Добавляем данные о ячейке, включая конфигурацию
                cell_data.append({
                    "cell_name": cell_name,
                    "value": cell_value,
                    "config": cell_config
                })

        # Печатаем собранные данные для отладки
        print("Собранные данные таблицы:", cell_data)
        return json.dumps(cell_data, ensure_ascii=False)

    @staticmethod
    def unmerge_cells(table, top_row, bottom_row, left_col, right_col):
        # Считываем текущее значение основной ячейки (левой верхней)
        main_item = table.item(top_row, left_col)
        if main_item:
            cell_data_str = main_item.data(Qt.UserRole)
            main_value = main_item.text() if not cell_data_str else json.loads(cell_data_str).get("value", "")
        else:
            main_value = ""

        # Отменяем объединение ячеек
        table.setSpan(top_row, left_col, 1, 1)

        # Восстанавливаем каждую ячейку в диапазоне
        for row in range(top_row, bottom_row + 1):
            for col in range(left_col, right_col + 1):
                if row == top_row and col == left_col:
                    # Левой верхней ячейке присваиваем основное значение
                    new_item = QTableWidgetItem(main_value)
                    config = {"merger": None}  # Обновляем конфигурацию
                    new_item.setData(Qt.UserRole, json.dumps(config))
                else:
                    # Остальные ячейки делаем пустыми
                    new_item = QTableWidgetItem("")
                    config = {"merger": None}
                    new_item.setData(Qt.UserRole, json.dumps(config))

                table.setItem(row, col, new_item)

    @staticmethod
    def is_merged(table, top_row, left_col):
        item = table.item(top_row, left_col)
        if item:
            cell_data = item.data(Qt.UserRole)
            if cell_data:
                try:
                    cell_data_json = json.loads(cell_data)
                    return cell_data_json.get("merger", "")
                except json.JSONDecodeError:
                    return False
        return False

    @staticmethod
    def refresh_table_view(table, template_data):
        """Обновляет таблицу с данными из шаблона, используя JSON."""
        try:
            # Сброс объединений и очистка таблицы
            TemplateTableService.reset_table(table)

            # Получаем основные параметры шаблона и устанавливаем структуру таблицы
            row_count = template_data.get("row_count", 0)
            col_count = template_data.get("col_count", 0)
            background_color = template_data.get("background_color", "#FFFFFF")
            cells = template_data.get("cells", [])

            TemplateTableService.setup_table_structure(table, row_count, col_count, background_color)

            # Обработка данных ячеек
            merged_cells = TemplateTableService.populate_table_cells(table, cells)

            # Обработка объединенных ячеек
            TemplateTableService.apply_merged_cells(table, merged_cells)

        except Exception as e:
            print(f"Error refreshing table view: {e}")

    @staticmethod
    def reset_table(table):
        """Сбрасывает объединения и очищает таблицу."""
        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                table.setSpan(row, col, 1, 1)  # Сбрасываем объединение для каждой ячейки
        table.clearContents()

    @staticmethod
    def setup_table_structure(table, row_count, col_count, background_color):
        """Устанавливает количество строк и столбцов в таблице и обновляет цвет фона."""
        table.setRowCount(row_count)
        table.setColumnCount(col_count)
        column_labels = [TemplateTableService.generate_col_name(col) for col in range(col_count)]
        table.setHorizontalHeaderLabels(column_labels)
        table.setVerticalHeaderLabels([str(i + 1) for i in range(row_count)])
        table.setStyleSheet(f"background-color: {background_color};")

    @staticmethod
    def populate_table_cells(table, cells):
        """Заполняет таблицу данными ячеек и сохраняет информацию об объединенных ячейках."""
        merged_cells = []

        for cell in cells:
            cell_name = cell.get("cell_name")
            value = cell.get("value")
            config = cell.get("config", {})

            # Преобразуем имя ячейки в индексы строки и столбца
            col_label = ''.join(filter(str.isalpha, cell_name))
            row_label = ''.join(filter(str.isdigit, cell_name))
            col_index = TemplateTableService.column_label_to_index(col_label)
            row_index = int(row_label) - 1

            # Создаем элемент для таблицы
            item = QTableWidgetItem()
            item.setText(value if value is not None else "")  # Устанавливаем значение, если оно не равно None

            # Устанавливаем конфигурации ячейки
            TemplateTableService.apply_cell_configuration(item, table, row_index, col_index, config)
            # Сохраняем всю конфигурацию в Qt.UserRole
            TemplateTableService.set_cell_data(item, value, config)

            table.setItem(row_index, col_index, item)

            # Проверяем, является ли ячейка объединенной, и сохраняем диапазон
            merge_range = config.get("merger")
            if merge_range:
                merged_cells.append(merge_range)

        return merged_cells

    @staticmethod
    def apply_cell_configuration(item, table, row_index, col_index, config):
        """Применяет конфигурации к ячейке, такие как шрифт, цвет, размеры и т.д."""
        # Цвет ячейки
        background_color = config.get("background_color")
        if background_color:
            item.setBackground(QtGui.QColor(background_color))

        # Цвет текста
        text_color = config.get("text_color")
        if text_color:
            item.setForeground(QtGui.QColor(text_color))

        # Шрифт текста
        font = config.get("font")
        if font:
            q_font = QtGui.QFont()
            q_font.fromString(font)
            item.setFont(q_font)

        # Наклон текста
        if config.get("text_tilt"):
            font = item.font()
            font.setItalic(True)
            item.setFont(font)

        # Подчеркивание текста
        if config.get("underline"):
            font = item.font()
            font.setUnderline(True)
            item.setFont(font)

        # Жирный текст
        if config.get("bold"):
            font = item.font()
            font.setBold(True)
            item.setFont(font)

        # Размер текста
        text_size = config.get("text_size")
        if text_size:
            font = item.font()
            font.setPointSize(text_size)
            item.setFont(font)

        # Применение форматирования чисел, если указано
        format_decimals = config.get("format")
        if format_decimals is not None:
            try:
                # Преобразуем текст в число и форматируем
                numeric_value = float(item.text())
                formatted_value = f"{numeric_value:.{format_decimals}f}"
                item.setText(formatted_value)
            except ValueError:
                # Если значение не является числом, оставляем текст без изменений
                pass

        # Высота и ширина ячейки
        height = config.get("height")
        width = config.get("width")
        # Устанавливаем высоту строки
        if height:
            table.setRowHeight(row_index, height)
        else:
            table.setRowHeight(row_index, 20)  # Дефолтное значение высоты строки

        # Устанавливаем ширину столбца
        if width:
            table.setColumnWidth(col_index, width)
        else:
            table.setColumnWidth(col_index, 100)  # Дефолтное значение ширины столбца

        # Выравнивание текста (добавлено)
        alignment = config.get("alignment", Qt.AlignLeft)
        item.setTextAlignment(alignment)

    @staticmethod
    def apply_merged_cells(table, merged_cells):
        """Применяет объединения ячеек к таблице."""
        for merge_range in merged_cells:
            top_row, left_col, bottom_row, right_col = TemplateTableService.parse_merge_range(merge_range)
            table.setSpan(top_row, left_col, (bottom_row - top_row + 1), (right_col - left_col + 1))

    @staticmethod
    def set_cell_data(item, value, config):
        """Устанавливает данные в QTableWidgetItem."""
        # Добавляем значение в конфигурацию, если оно отсутствует
        config["value"] = value
        item.setData(Qt.UserRole, json.dumps(config))
        item.setText(value)
        item.setTextAlignment(Qt.AlignCenter)  # Центрируем текст в ячейке

    @staticmethod
    def parse_merge_range(merge_range):
        """Разбирает строку диапазона объединения (например, 'C1:D2') и возвращает индексы начальной и конечной ячейки."""
        try:
            top_left, bottom_right = merge_range.split(":")
            top_row, left_col = TemplateTableService.parse_cell_position(top_left)
            bottom_row, right_col = TemplateTableService.parse_cell_position(bottom_right)
            return top_row, left_col, bottom_row, right_col
        except ValueError:
            print(f"Ошибка парсинга диапазона объединения: {merge_range}")
            return 0, 0, 0, 0

    @staticmethod
    def parse_cell_position(cell_name):
        """Разбирает имя ячейки (например, 'A1') и возвращает индексы строки и столбца."""
        col_label = ''.join(filter(str.isalpha, cell_name))
        row_label = ''.join(filter(str.isdigit, cell_name))
        col_index = TemplateTableService.column_label_to_index(col_label)
        row_index = int(row_label) - 1
        return row_index, col_index

    @staticmethod
    def toggle_bold(table):
        """Переключает жирность текста для выделенных ячеек."""
        selected_ranges = table.selectedRanges()
        for selected_range in selected_ranges:
            for row in range(selected_range.topRow(), selected_range.bottomRow() + 1):
                for col in range(selected_range.leftColumn(), selected_range.rightColumn() + 1):
                    item = table.item(row, col)
                    if item is None:
                        item = QTableWidgetItem()
                        table.setItem(row, col, item)

                    # Получаем текущую конфигурацию ячейки
                    cell_data_str = item.data(Qt.UserRole)
                    cell_data = json.loads(cell_data_str) if cell_data_str else {}

                    # Переключаем жирность текста
                    current_font = item.font()
                    is_bold = not current_font.bold()  # Инвертируем текущее состояние жирности
                    current_font.setBold(is_bold)
                    item.setFont(current_font)

                    # Обновляем конфигурацию
                    if is_bold:
                        cell_data['bold'] = True
                    else:
                        cell_data.pop('bold', None)  # Убираем из конфигурации

                    # Сохраняем данные в Qt.UserRole
                    item.setData(Qt.UserRole, json.dumps(cell_data))

    @staticmethod
    def toggle_italic(table):
        """Переключает наклон текста для выделенных ячеек."""
        selected_ranges = table.selectedRanges()
        for selected_range in selected_ranges:
            for row in range(selected_range.topRow(), selected_range.bottomRow() + 1):
                for col in range(selected_range.leftColumn(), selected_range.rightColumn() + 1):
                    item = table.item(row, col)
                    if item is None:
                        item = QTableWidgetItem()
                        table.setItem(row, col, item)

                    # Получаем текущую конфигурацию ячейки
                    cell_data_str = item.data(Qt.UserRole)
                    cell_data = json.loads(cell_data_str) if cell_data_str else {}

                    # Переключаем наклон текста
                    current_font = item.font()
                    is_italic = not current_font.italic()  # Инвертируем текущее состояние наклона
                    current_font.setItalic(is_italic)
                    item.setFont(current_font)

                    # Обновляем конфигурацию
                    if is_italic:
                        cell_data['text_tilt'] = True
                    else:
                        cell_data.pop('text_tilt', None)  # Убираем из конфигурации

                    # Сохраняем данные в Qt.UserRole
                    item.setData(Qt.UserRole, json.dumps(cell_data))

    @staticmethod
    def toggle_underline(table):
        """Переключает подчеркивание текста для выделенных ячеек."""
        selected_ranges = table.selectedRanges()
        for selected_range in selected_ranges:
            for row in range(selected_range.topRow(), selected_range.bottomRow() + 1):
                for col in range(selected_range.leftColumn(), selected_range.rightColumn() + 1):
                    item = table.item(row, col)
                    if item is None:
                        item = QTableWidgetItem()
                        table.setItem(row, col, item)

                    # Получаем текущую конфигурацию ячейки
                    cell_data_str = item.data(Qt.UserRole)
                    cell_data = json.loads(cell_data_str) if cell_data_str else {}

                    # Переключаем подчеркивание текста
                    current_font = item.font()
                    is_underline = not current_font.underline()  # Инвертируем текущее состояние подчеркивания
                    current_font.setUnderline(is_underline)
                    item.setFont(current_font)

                    # Обновляем конфигурацию
                    if is_underline:
                        cell_data['underline'] = True
                    else:
                        cell_data.pop('underline', None)  # Убираем из конфигурации

                    # Сохраняем данные в Qt.UserRole
                    item.setData(Qt.UserRole, json.dumps(cell_data))

    @staticmethod
    def apply_text_color(table, color):
        """Применяет цвет текста к выделенным ячейкам."""
        selected_ranges = table.selectedRanges()
        for selected_range in selected_ranges:
            for row in range(selected_range.topRow(), selected_range.bottomRow() + 1):
                for col in range(selected_range.leftColumn(), selected_range.rightColumn() + 1):
                    item = table.item(row, col)
                    if item is None:
                        item = QTableWidgetItem()
                        table.setItem(row, col, item)

                    # Устанавливаем цвет текста
                    item.setForeground(QColor(color))

                    # Сохраняем цвет текста в конфигурации
                    cell_data_str = item.data(Qt.UserRole)
                    if cell_data_str:
                        cell_data = json.loads(cell_data_str)
                    else:
                        cell_data = {}

                    cell_data['text_color'] = color
                    item.setData(Qt.UserRole, json.dumps(cell_data))

    @staticmethod
    def apply_cell_background_color(table, color):
        """Применяет цвет фона к выделенным ячейкам."""
        selected_ranges = table.selectedRanges()
        for selected_range in selected_ranges:
            for row in range(selected_range.topRow(), selected_range.bottomRow() + 1):
                for col in range(selected_range.leftColumn(), selected_range.rightColumn() + 1):
                    item = table.item(row, col)
                    if item is None:
                        item = QTableWidgetItem()
                        table.setItem(row, col, item)

                    # Устанавливаем цвет фона
                    item.setBackground(QColor(color))

                    # Сохраняем цвет фона в конфигурации
                    cell_data_str = item.data(Qt.UserRole)
                    if cell_data_str:
                        cell_data = json.loads(cell_data_str)
                    else:
                        cell_data = {}

                    cell_data['background_color'] = color
                    item.setData(Qt.UserRole, json.dumps(cell_data))

    # @staticmethod
    # def change_decimal_places(table, increase=True):
    #     """Изменяет количество знаков после запятой для выбранных ячеек."""
    #     selected_ranges = table.selectedRanges()
    #
    #
    #     for selected_range in selected_ranges:
    #         for row in range(selected_range.topRow(), selected_range.bottomRow() + 1):
    #             for col in range(selected_range.leftColumn(), selected_range.rightColumn() + 1):
    #                 item = table.item(row, col)
    #                 if item is None:
    #                     item = QTableWidgetItem()
    #                     table.setItem(row, col, item)
    #
    #                 # Получаем текущую конфигурацию ячейки
    #                 cell_data_str = item.data(Qt.UserRole)
    #                 if cell_data_str:
    #                     cell_config = json.loads(cell_data_str)
    #                 else:
    #                     cell_config = {}
    #
    #                 # Изменяем количество знаков после запятой
    #                 current_format = cell_config.get("format", 0)
    #                 if current_format is None:
    #                     current_format = 0
    #                 if increase:
    #                     current_format += 1
    #                 else:
    #                     current_format = max(0, current_format - 1)
    #
    #                 # Сохраняем изменённый формат
    #                 cell_config["format"] = current_format
    #                 item.setData(Qt.UserRole, json.dumps(cell_config))
    #
    #                 # Если значение в ячейке является числом, обновляем отображение
    #                 value = item.text()
    #                 try:
    #                     numeric_value = float(value)
    #                     formatted_value = f"{numeric_value:.{current_format}f}"
    #                     item.setText(formatted_value)
    #                 except ValueError:
    #                     # Если значение не число, оставляем текст как есть
    #                     pass

    @staticmethod
    def change_decimal_places(table, increase=True):
        """Изменяет количество знаков после запятой для выбранных ячеек."""
        selected_ranges = table.selectedRanges()

        for selected_range in selected_ranges:
            for row in range(selected_range.topRow(), selected_range.bottomRow() + 1):
                for col in range(selected_range.leftColumn(), selected_range.rightColumn() + 1):
                    item = table.item(row, col)
                    if item is None:
                        item = QTableWidgetItem()
                        table.setItem(row, col, item)

                    # Получаем текущую конфигурацию ячейки
                    cell_data_str = item.data(Qt.UserRole)
                    if cell_data_str:
                        cell_config = json.loads(cell_data_str)
                    else:
                        cell_config = {}

                    # Изменяем количество знаков после запятой
                    current_format = cell_config.get("format", 0)
                    if current_format is None:
                        current_format = 0
                    if increase:
                        current_format += 1
                    else:
                        current_format = max(0, current_format - 1)

                    # Сохраняем изменённый формат
                    cell_config["format"] = current_format
                    item.setData(Qt.UserRole, json.dumps(cell_config))

                    # Если значение в ячейке является числом, обновляем отображение
                    value = item.text()
                    try:
                        numeric_value = float(value)
                        formatted_value = f"{numeric_value:.{current_format}f}"
                        item.setText(formatted_value)
                    except ValueError:
                        # Если значение не число, оставляем текст как есть
                        pass


    @staticmethod
    def change_font_size(table, font_size):
        """Изменяет размер шрифта для выделенных ячеек."""
        selected_ranges = table.selectedRanges()

        for selected_range in selected_ranges:
            for row in range(selected_range.topRow(), selected_range.bottomRow() + 1):
                for col in range(selected_range.leftColumn(), selected_range.rightColumn() + 1):
                    item = table.item(row, col)
                    if item is None:
                        item = QTableWidgetItem()
                        table.setItem(row, col, item)

                    # Получаем текущую конфигурацию ячейки
                    cell_data_str = item.data(Qt.UserRole)
                    if cell_data_str:
                        cell_config = json.loads(cell_data_str)
                    else:
                        cell_config = {}

                    # Устанавливаем новый размер шрифта
                    cell_config["text_size"] = font_size

                    # Применяем новый шрифт
                    font = item.font()
                    font.setPointSize(font_size)
                    item.setFont(font)

                    # Сохраняем конфигурацию
                    item.setData(Qt.UserRole, json.dumps(cell_config))

                    # Центрируем текст в ячейке
                    item.setTextAlignment(Qt.AlignCenter)

    @staticmethod
    def resize_table(table, new_row_count, new_col_count, background_color):
        """Изменяет размер таблицы и сохраняет размеры строк и столбцов."""
        # Сохранение текущих размеров строк и столбцов
        row_sizes = {row: table.rowHeight(row) for row in range(table.rowCount())}
        col_sizes = {col: table.columnWidth(col) for col in range(table.columnCount())}

        # Изменение структуры таблицы
        table.setRowCount(new_row_count)
        table.setColumnCount(new_col_count)
        table.setStyleSheet(f"background-color: {background_color};")

        # Восстановление размеров строк
        for row, height in row_sizes.items():
            if row < new_row_count:  # Применяем только для существующих строк
                table.setRowHeight(row, height)

        # Восстановление размеров столбцов
        for col, width in col_sizes.items():
            if col < new_col_count:  # Применяем только для существующих столбцов
                table.setColumnWidth(col, width)

    @staticmethod
    def populate_excel_with_data(ws, table_widget):
        """Заполняет Excel-таблицу данными, форматами и размерами."""
        for row in range(table_widget.rowCount()):
            for col in range(table_widget.columnCount()):
                item = table_widget.item(row, col)
                if item:
                    # Write the text value to the cell
                    text = item.text()
                    cell = ws.cell(row=row + 1, column=col + 1, value=text)

                    # Apply cell formatting (font, alignment, etc.)
                    TemplateTableService.apply_cell_formatting(cell, item)

                    # Handle cell dimensions (width and height)
                    TemplateTableService.handle_cell_dimensions(ws, row, col, table_widget)

    @staticmethod
    def apply_cell_formatting(cell, item):
        """
        Применяет форматирование и размеры к ячейке Excel на основе данных из Qt.UserRole.
        Если данные отсутствуют, применяются стандартные значения.
        """
        cell_data_str = item.data(Qt.UserRole)
        if cell_data_str:
            try:
                cell_data = json.loads(cell_data_str)

                # Устанавливаем шрифт (по умолчанию "Calibri", размер 12)
                font_string = cell_data.get("font", "Calibri,12")
                if isinstance(font_string, str):
                    font_parts = font_string.split(",")
                    font_name = font_parts[0] if len(font_parts) > 0 else "Calibri"
                else:
                    font_name = "Calibri"
                font_size = cell_data.get("text_size", 12)
                bold = cell_data.get("bold", False) if cell_data.get("bold") is not None else False
                italic = cell_data.get("text_tilt", False) if cell_data.get("text_tilt") is not None else False
                underline = "single" if cell_data.get("underline", False) else None

                # Устанавливаем цвет текста (по умолчанию чёрный)
                text_color = cell_data.get("text_color", "#000000")
                text_color = text_color.lstrip("#") if isinstance(text_color, str) and text_color.startswith(
                    "#") else "000000"
                # Устанавливаем цвет фона (по умолчанию белый)
                background_color = cell_data.get("background_color", "#FFFFFF")
                background_color = background_color.lstrip("#") if isinstance(background_color,
                                                                              str) and background_color.startswith(
                    "#") else "FFFFFF"

                # Применяем форматирование текста
                cell.font = Font(
                    name=font_name,
                    size=font_size,
                    bold=bold,
                    italic=italic,
                    underline=underline,
                    color=text_color,
                )

                # Устанавливаем цвет фона ячейки
                cell.fill = PatternFill(
                    start_color=background_color,
                    end_color=background_color,
                    fill_type="solid",
                )

                # Устанавливаем выравнивание текста (по умолчанию по центру)
                alignment = cell_data.get("alignment", {})
                cell.alignment = Alignment(
                    horizontal=alignment.get("horizontal", "center"),
                    vertical=alignment.get("vertical", "center"),
                )

                # Обрабатываем размеры ячейки
                width = cell_data.get("width", 100)  # Ширина по умолчанию
                height = cell_data.get("height", 20)  # Высота по умолчанию

                col_letter = get_column_letter(cell.column)

                # Устанавливаем ширину столбца
                cell.parent.column_dimensions[col_letter].width = max(float(width) / 7,
                                                                      10)  # Excel использует ширину в условных единицах
                cell.parent.row_dimensions[cell.row].height = height

                # Устанавливаем границы (по умолчанию чёрные тонкие линии)
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            except (json.JSONDecodeError, ValueError, TypeError) as e:
                print(f"Ошибка обработки ячейки: {e}")
                print(f"Данные ячейки: {cell_data_str}")

    @staticmethod
    def handle_cell_dimensions(ws, row, col, table_widget):
        """Обрабатывает размеры строк и столбцов."""
        # Получаем ширину и высоту из данных ячейки
        item = table_widget.item(row, col)
        if item:
            cell_data_str = item.data(Qt.UserRole)
            if cell_data_str:
                try:
                    cell_data = json.loads(cell_data_str)
                    width = cell_data.get("width", table_widget.columnWidth(col))
                    height = cell_data.get("height", table_widget.rowHeight(row))

                    if width:
                        col_letter = get_column_letter(col + 1)
                        ws.column_dimensions[col_letter].width = max(float(width) / 7, 10)
                    if height:
                        ws.row_dimensions[row + 1].height = max(float(height) / 15, 15)
                except json.JSONDecodeError:
                    print(f"Ошибка декодирования JSON для размеров ячейки: {cell_data_str}")

    @staticmethod
    def handle_merged_cells(ws, table_widget):
        """Обрабатывает объединенные ячейки в таблице."""
        for row in range(table_widget.rowCount()):
            for col in range(table_widget.columnCount()):
                item = table_widget.item(row, col)
                if item:
                    cell_data_str = item.data(Qt.UserRole)
                    if cell_data_str:
                        try:
                            cell_data = json.loads(cell_data_str)
                            if "merger" in cell_data and cell_data["merger"]:
                                merger = cell_data["merger"]  # Формат: "A1:C1"
                                start_cell, end_cell = merger.split(":")
                                ws.merge_cells(f"{start_cell}:{end_cell}")
                        except json.JSONDecodeError:
                            print(f"Ошибка обработки объединённых ячеек ({row}, {col})")

    @staticmethod
    def export_to_excel(table_widget):
        """
        Экспорт данных таблицы с конфигурациями в файл Excel.
        :param table_widget: QTableWidget - виджет с данными таблицы
        """
        # Выбираем путь для сохранения файла
        file_path, _ = QFileDialog.getSaveFileName(None, "Сохранить как", "", "Excel Files (*.xlsx)")
        if not file_path:
            return  # Пользователь отменил выбор

        try:
            # Создаем новую книгу Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Экспорт таблицы"

            # Populate data, formats, and dimensions
            TemplateTableService.populate_excel_with_data(ws, table_widget)

            # Handle merged cells
            TemplateTableService.handle_merged_cells(ws, table_widget)

            # Сохраняем файл
            wb.save(file_path)
            QMessageBox.information(None, "Успех", f"Файл успешно сохранен в {file_path}")

        except Exception as e:
            QMessageBox.critical(None, "Ошибка", f"Не удалось сохранить файл: {str(e)}")


